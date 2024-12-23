import secrets
import logging
import concurrent.futures
from io import BytesIO

from werkzeug.datastructures import FileStorage
from azure.core.exceptions import ResourceExistsError, ResourceNotFoundError
import openpyxl as oxl

from data_stores.AzureTableObjects import AzureTableObjects as ato
from data_stores.AzureBlobObjects import AzureBlobObjects as abo
from initializers import deletingFuncs as delf
from initializers import extract_data_for_review
from data_stores.AzureTableObjects import MutexError
from initializers.extract_data_to_json_store import extract_data_to_json_store
from initializers.create_vectorstore import update_vectorstore
import config

def allowed_file(filename: str, extension: str) -> bool:
    """Checks if filename has allowed extension.
    Example: filename = '28.pdf', extension = 'pdf' will return true
    """
    actualExtension = filename.rsplit(".")[-1]
    return actualExtension == extension

def validateUpload(extension: str, file: FileStorage) -> str:
    """Checks if the document POSTed by the user matches the extension

    Args:
        extension (str): expected extension of the file user is expected to POST. eg: 'pdf'

    Returns:
        str: errors. Blank string if no error
    """
    
    errors = ''
    
    if file and not allowed_file(file.filename, extension):
        errors = f'File does not match extension {extension}'

    return errors

def generateArrayForTableRows() -> list[list[str]]:
    """Generates the list required to construct the dynamic table in the file management interface.
    The list contains lists representing items in a single table row (i.e. file names, and status)

    Returns:
        list[list[str]]: table rows
    """
    tableRows: list[list[str]] = []
    # listOfPDFNames = abo.getListOfFilenamesInContainer(config.pdf_container_name) # no need bcuz there if a record exists for a chapter, it will definitely have a PDF
    listOfGeneratedExcelNames = abo.getListOfFilenamesInContainer(config.generatedExcel_container_name)
    listOfReviewedExcelNames = abo.getListOfFilenamesInContainer(config.reviewedExcel_container_name)
    listOfJSONs = abo.getListOfFilenamesInContainer(config.json_container_name)
    entities = ato.get_all_chapter_records()
    for entity in entities:
        release_date = entity['RowKey'].rsplit(':')[0]
        chapterNumber = entity['RowKey'].rsplit(':')[1]
        excelName = chapterNumber + '.xlsx'
        jsonName = chapterNumber + '.json'
        tableRow = [release_date,chapterNumber,f'{chapterNumber}.pdf']
        if f'{release_date}/{excelName}' in listOfGeneratedExcelNames: tableRow.append(excelName)
        else: tableRow.append('Nil')
        if f'{release_date}/{excelName}' in listOfReviewedExcelNames: tableRow.append(excelName)
        else: tableRow.append('Nil')
        if f'{release_date}/{jsonName}' in listOfJSONs: tableRow.append(jsonName)
        else: tableRow.append('Nil')
        status = entity['RecordStatus']
        tableRow += [status]
        tableRows.append(tableRow)
    return tableRows

def delete_upto_corrected_excel(chapterNumber: int, release: str):
    """Deletes a file record upto and including the corrected excel.

    """
    job_id = ato.create_new_job('delete_upto_corrected_excel', f'chapterNumber: {chapterNumber}, release: {release}')

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(delf.deleteChapterFromCosmos, chapterNumber, release), 
            executor.submit(delf.deleteChapterJsonBlob, chapterNumber, release), 
            executor.submit(delf.deleteChapterReviewedExcelBlob, chapterNumber, release)
        ]
        concurrent.futures.wait(futures)

    ato.set_job_progress(job_id, 'done')
    ato.end_job(job_id)
    
def delete_upto_pdf(chapterNumber: int, release: str):
    """Deletes the entire file record.

    """
    job_id = ato.create_new_job('delete_upto_pdf', f'chapterNumber: {chapterNumber}, release: {release}')

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(delete_upto_corrected_excel, chapterNumber, release),
            executor.submit(delf.deleteChapterDictPickleBlob, chapterNumber, release),
            executor.submit(delf.deleteChapterGeneratedExcelBlob, chapterNumber, release),
            executor.submit(delf.deleteChapterPDFBlob, chapterNumber, release)
        ]
        concurrent.futures.wait(futures)
    
    ato.set_job_progress(job_id, 'done')
    ato.end_job(job_id)
    
def upload_pdf(pdffile: BytesIO, release_date: str,user_entered_chapter_number: int = None, filename: str = None, is_called_by_batch: bool = False):
    """Basically does the entire uploadPDF stage mentioned in section 3.1.3 in the developer guide.
    is_called_by_batch argument is used if the function is calle by batch_upload_pdf - if False, this function will handle job tracking"""
    
    if not is_called_by_batch:
        job_id = ato.create_new_job('upload_pdf', f'release_date: {release_date}, user_entered_chapter_number: {user_entered_chapter_number}, filename: {filename}')

    # convert the PDF into the dictionary, excel and identified chapter number
    dictionary_pkl_stream, excel_stream, chapterNumber = extract_data_for_review.convertPDFToExcelForReview(pdffile,user_entered_chapter_number,filename)
    if not dictionary_pkl_stream: # i.e. an exception was raised when trying to extract data from the pdf
        if user_entered_chapter_number:
            logging.error(f'Error with pdf or entered chapter number. User entered chapter number: {user_entered_chapter_number}. Release {release_date}')
        elif filename:
            logging.error(f"Error with pdf or entered chapter number. User uploaded file's filename: {filename}. Release {release_date}")
        elif chapterNumber:
            logging.error(f"Error with pdf or entered chapter number. Filename and user_entered_chapter_number not received. Identified chapter number: {str(chapterNumber)}. Release {release_date}")
        if not is_called_by_batch:
            ato.set_job_progress(job_id, 'error with pdf')
            ato.end_job(job_id)
        return
    
    
    try: ato.create_new_blank_chapter_record(chapterNumber, release_date)
    except ResourceExistsError: 
        logging.error(f'A record for chapter {chapterNumber} (release {release_date}) already exists. Delete it if you want to upload a new PDF.')
        if not is_called_by_batch:
            ato.set_job_progress(job_id, 'record already existed')
            ato.end_job(job_id)
        return

    mutexKey = secrets.token_hex()
    try: ato.claim_mutex(chapterNumber, mutexKey, release_date)
    except MutexError as e:
        logging.error(e.__str__())
        if not is_called_by_batch: ato.end_job(job_id)
        return
    ato.edit_chapter_record(chapterNumber, mutexKey, release_date, newRecordStatus=config.RecordStatus.uploadingPDF)
    pdffile.seek(0)
    abo.upload_to_blob_from_stream(pdffile, config.pdf_container_name, f'{release_date}/{chapterNumber}.pdf') # PDF uploaded to azure blob
    logging.log(25,f'{chapterNumber}.pdf' + f' of release {release_date} ' + ' successfully uploaded')

    ato.edit_chapter_record(chapterNumber, mutexKey, release_date, newRecordStatus=config.RecordStatus.uploadingGeneratedDocuments)
    abo.upload_to_blob_from_stream(dictionary_pkl_stream, config.generatedDict_container_name, f'{release_date}/{chapterNumber}.pkl') # upload generated excel to azure blob
    abo.upload_to_blob_from_stream(excel_stream, config.generatedExcel_container_name, f'{release_date}/{chapterNumber}.xlsx') # upload dictionary pickle to azure blob
    ato.edit_chapter_record(chapterNumber, mutexKey, release_date, newRecordStatus='', newRecordState=config.RecordState.pdfUploaded)
    ato.release_mutex(chapterNumber, mutexKey, release_date)
      
    logging.info("Data extracted from tariff pdfs and saved as excel (and text data dictionary pickle) for review.")
    if not is_called_by_batch:
        ato.set_job_progress(job_id, 'done')
        ato.end_job(job_id)

def batch_upload_pdfs(pdffiles: list[BytesIO], release_date: str, filenames: list[str] = None):
    job_description = f'Release Date: {release_date} filenames: {filenames}'
    job_id = ato.create_new_job('batch_upload_pdfs', job_description)
    for i,pdffile in enumerate(pdffiles): # uploads happen in series, so time taken for the total upload process is the same, but memory is saved
        filename = filenames[i]
        upload_pdf(pdffile,release_date,filename=filename, is_called_by_batch=True)
        current_progress = ato.get_job_progress(job_id)
        current_progress += filename + ','
        ato.set_job_progress(job_id, current_progress)
    current_progress = ato.get_job_progress(job_id)
    if len(current_progress) > 0: current_progress = current_progress[:-1] # remove final ','
    ato.set_job_progress(job_id, current_progress)
    ato.end_job(job_id)

def upload_excel(excelfile: BytesIO, filename: str, release_date: str, user_entered_chapter_number: int = None, is_called_by_batch: bool = False):
    """Basically does the entire uploadExcel stage mentioned in section 3.1.3 in the developer guide."""
    
    if not is_called_by_batch:
        job_id = ato.create_new_job('upload_excel', f'release_date: {release_date}, user_entered_chapter_number: {user_entered_chapter_number}, filename: {filename}')

    # if user has entered the chapter number, that is taken as the chapterNumber, otherwise it's taken from the filename
    if user_entered_chapter_number: chapterNumber = user_entered_chapter_number
    else: 
        chapterNumber = filename.rsplit('.')[0]
        try: chapterNumber = int(chapterNumber)
        except ValueError: 
            logging.error(f'Cannot identify the chapter number the excel {filename} of release {release_date} refers to')
            if not is_called_by_batch:
                ato.set_job_progress(job_id,'cannot identify the chapter number the excel belongs to')
                ato.end_job(job_id)
            return

    mutexKey = secrets.token_hex()
    try: ato.claim_mutex(chapterNumber, mutexKey, release_date)
    except MutexError as e:
        logging.error(e.__str__())
        if not is_called_by_batch:
            ato.end_job(job_id)
        return
    except ResourceNotFoundError:
        logging.error(f'The chapter was not found. Perhaps you must create the chapter record by uploading a PDF. Excel: {filename}. Release: {release_date}')
        if not is_called_by_batch:
            ato.set_job_progress(job_id,'chapter record was not found')
            ato.end_job(job_id)
        return

    isSuccess = extract_data_to_json_store(excelfile, mutexKey, chapterNumber, release_date)
    if isSuccess:
        logging.log(25,'Excel and generated json successfully uploaded.')
    else:
        logging.error(f'Excel was rejected due to an error. Maybe at least one of the HS codes provided did not match the entered chapter number. Excel: {filename}. Release: {release_date}')
        ato.release_mutex(chapterNumber, mutexKey, release_date)
        if not is_called_by_batch:
            ato.set_job_progress(job_id,'excel was rejected due to an error')
            ato.end_job(job_id)
        return
    
    update_vectorstore(chapterNumber, mutexKey, release_date)

    if not is_called_by_batch:
        ato.set_job_progress(job_id, 'done')
        ato.end_job(job_id)

def batch_upload_excels(excelfiles: list[BytesIO], release_date: str, filenames: list[str] = None):
    job_description = f'Release Date: {release_date} filenames: {filenames}'
    job_id = ato.create_new_job('batch_upload_excels', job_description)
    for i,excelfile in enumerate(excelfiles): # uploads happen in series, so time taken for the total upload process is the same, but memory is saved
        filename = filenames[i]
        upload_excel(excelfile,filename,release_date, is_called_by_batch=True)
        current_progress = ato.get_job_progress(job_id)
        current_progress += filename + ','
        ato.set_job_progress(job_id, current_progress)
    current_progress = ato.get_job_progress(job_id)
    if len(current_progress) > 0: current_progress = current_progress[:-1] # remove final ','
    ato.set_job_progress(job_id, current_progress)
    ato.end_job(job_id)
    

def add_release(release: str):
    """Used to declare a date-string as a release.
    The functions add_release, remove_release and get_stored_releases can be used by front-end to offer dropdown menus for user to select release date.
    Currently not used to enforce anything.
    This system is used to make radio or dropdown buttons for filters in some html pages.
    """
    existing_streamed = abo.download_blob_file_to_stream(config.release_holder_filename, config.release_holder_container_name)
    existing_streamed.seek(0)
    existing_text = existing_streamed.getvalue().decode('utf-8')
    existing_releases = existing_text.rsplit('\n')
    for existing_release in existing_releases:
        _existing_release = existing_release.rstrip()
        _release = release.rstrip()
        if _existing_release == _release: return
    new_text = existing_text + release + '\n'
    new_streamed = BytesIO(new_text.encode('utf-8'))
    new_streamed.seek(0)
    abo.upload_to_blob_from_stream(new_streamed,config.release_holder_container_name,config.release_holder_filename)

def remove_release(release: str):
    """Used to remove a release.
    The functions add_release, remove_release and get_stored_releases can be used by front-end to offer dropdown menus for user to select release date.
    Currently not used to enforce anything.
    This system is used to make radio or dropdown buttons for filters in some html pages.
    """
    existing_releases = get_stored_releases()
    for i in range(0,len(existing_releases)):
        existing_release = existing_releases[i]
        _release = release.strip()
        _existing_release = existing_release.strip()
        if _release == _existing_release:
            logging.info(f'Release {release} removed.') 
            existing_releases.pop(i)
            break
    new_text = ''
    for each in existing_releases:
        if each.strip() == '': continue
        new_text += each + '\n'
    new_streamed = BytesIO(new_text.encode('utf-8'))
    new_streamed.seek(0)
    abo.upload_to_blob_from_stream(new_streamed,config.release_holder_container_name,config.release_holder_filename)

def get_stored_releases() -> list[str]:
    """Used to get declared releases.
    The functions add_release, remove_release and get_stored_releases can be used by front-end to offer dropdown menus for user to select release date.
    Currently not used to enforce anything.
    This system is used to make radio or dropdown buttons for filters in some html pages.
    """
    existing_streamed = abo.download_blob_file_to_stream(config.release_holder_filename, config.release_holder_container_name)
    existing_streamed.seek(0)
    existing_text = existing_streamed.getvalue().decode('utf-8')
    existing_releases = existing_text.rsplit('\n')
    existing_releases=existing_releases[:-1] # last line is blank
    return existing_releases

def does_excel_have_hscodeerrors(excel: BytesIO) -> bool:
    """Checks if a given excel (typically the generated excel) has hscode errors that the user must review."""
    workbook = oxl.load_workbook(excel)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.max_column, max_col=sheet.max_column): # iterate over the last column
        for cell in row:
            content = cell.value
            if content == None: continue
            if content.rstrip() == 'hscode error': return True
    return False


def is_POST_data_empty(args: list):
    for arg in args:
        if arg == None or arg == '': return True
    return False
