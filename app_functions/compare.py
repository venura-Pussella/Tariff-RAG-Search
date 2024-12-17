import json
from io import BytesIO

import pandas as pd
import openpyxl as oxl
import openpyxl.styles as oxlstyles

import config
from data_stores.AzureBlobObjects import AzureBlobObjects as abo
from app_functions import findByHSCode as fhs


def compare_releases(chapter_number: int, release1: str, release2: str) -> tuple[list,list,list,list]:
    """Does git-like version comparison between two documents that is the same chapter but of 2 releases.
    Git compares changes in lines. This compares changes in HS Codes.
    IMPORTANT: Assumes HS codes are sorted in the chapter-release JSON.

    Args:
        chapter_number (int): _description_
        release1 (str): old release
        release2 (str): new release

    Returns:
        tuple[list,list,list,list]: hscodes_with_no_change,hscodes_with_change,new_hscodes,removed_hscodes
    """

    # download jsons
    release1_json_bin = abo.download_blob_file_to_stream(f'{release1}/{chapter_number}.json', config.json_container_name)
    release2_json_bin = abo.download_blob_file_to_stream(f'{release2}/{chapter_number}.json', config.json_container_name)
    release1_json_bin.seek(0); release2_json_bin.seek(0)

    # convert to string
    release1_json_string = release1_json_bin.getvalue().decode('utf-8')
    release2_json_string = release2_json_bin.getvalue().decode('utf-8')

    # convert to lists of dictionary
    release1_dict = json.loads(release1_json_string)
    release2_dict = json.loads(release2_json_string)

    # get items of each release
    release1_item_list: list[dict] = release1_dict['Items']
    release2_item_list: list[dict] = release2_dict['Items']

    # get changes (similar to git version control comparison)
    return __compare_item_lists_twoPointerMethod(release1_item_list, release2_item_list)


def __compare_item_lists_twoPointerMethod(list1: list[dict], list2: list[dict]):

    hscodes_with_no_change = []
    hscodes_with_change = []
    new_hscodes = []
    removed_hscodes = []

    i = 0; j = 0
    while True:
        left = list1[i]
        right = list2[j]

        if left['HS Code'] == right['HS Code']: # can make comparison
            if __are_items_equal(left, right): 
                hscodes_with_no_change.append(left['HS Code'])
                print(f"Added {left['HS Code']} to no change.")
            else: 
                hscodes_with_change.append(left['HS Code'])
                print(f"Added {left['HS Code']} to changed.")
            i += 1; j += 1

        if left['HS Code'] > right['HS Code']:
            new_hscodes.append(right['HS Code'])
            print(f"Added {right['HS Code']} to new.")
            j += 1

        if left['HS Code'] < right['HS Code']:
            removed_hscodes.append(left['HS Code'])
            print(f"Added {left['HS Code']} to old.")
            i += 1


        if i >= len(list1):
            while j < len(list2):
                right = list2[j]
                new_hscodes.append(right['HS Code'])
                print(f"Added {right['HS Code']} to new.")
                j += 1
            break

        if j >= len(list2):
            while i < len(list1):
                left = list1[i]
                removed_hscodes.append(left['HS Code'])
                print(f"Added {left['HS Code']} to old.")
                i += 1
            break

    return hscodes_with_no_change,hscodes_with_change,new_hscodes,removed_hscodes


def __are_items_equal(item1: dict, item2: dict):
    for key in item1.keys():
        if item1[key] != item2[key]: return False
    return True

def get_lineitems_for_display_from_hscodes(hscodes: list[str], release: str):
    """Converts a list of hs codes into lineitems for displaying in the html dyanamic table.
    """
    results = []
    for hscode in hscodes:
        results += fhs.findByHSCode(hscode,[release])
    return results

def compare_releases_all_possible_chapters(release1: str, release2: str):
    """Does git-like version comparison between every 2 possible documents (chapters) between 2 releases.
    Git compares changes in lines. This compares changes in HS Codes.
    IMPORTANT: Assumes HS codes are sorted in the chapter-release JSON.

    Args:
        release1 (str): old release
        release2 (str): new release

    Returns:
        tuple[list,list,list,list]: hscodes_with_no_change,hscodes_with_change,new_hscodes,removed_hscodes
    """
    hscodes_with_no_change = []
    hscodes_with_change = []
    new_hscodes = []
    removed_hscodes = []

    # identify all chapters that can be compared (based on per release json availability in Azure blob)
    filenames = abo.getListOfFilenamesInContainer(config.json_container_name)
    release1_chapters = set()
    release2_chapters = set()
    for name in filenames:
        release = name.rsplit('/')[0]
        chapter_number = int(name.rsplit('/')[1].rsplit('.')[0])
        if release == release1: release1_chapters.add(chapter_number)
        if release == release2: release2_chapters.add(chapter_number)
    intersecting_chapters = release1_chapters.intersection(release2_chapters)

    # do the comparison
    for chapter_number in intersecting_chapters:
        results = compare_releases(chapter_number,release1,release2)
        hscodes_with_no_change += results[0]; hscodes_with_change += results[1]; new_hscodes += results[2]; removed_hscodes += results[3]
    
    # sort and return
    new_hscodes.sort(); removed_hscodes.sort(); hscodes_with_change.sort(); hscodes_with_no_change.sort()
    return hscodes_with_no_change,hscodes_with_change,new_hscodes,removed_hscodes

def export_release_comparison(release1: str, release2: str) -> tuple[BytesIO,BytesIO,BytesIO]:
    """Compares all available chapters present in both of the 2 specified releases in a git-like fasion, 
    and generates 3 excel files containing new HS codes, removed HS codes, and changed HS codes.
    Git compares changes in lines. This compares changes in HS Codes.
    IMPORTANT: Assumes HS codes are sorted in the chapter-release JSON.

    Args:
        release1 (str): _description_
        release2 (str): _description_

    Returns:
        tuple[BytesIO,BytesIO,BytesIO]: new_lineitems_excel, removed_lineitems_excel, changed_excel
    """
    hscodes_with_no_change,hscodes_with_change,new_hscodes,removed_hscodes = compare_releases_all_possible_chapters(release1, release2)

    new_hscodes_lineitems = get_lineitems_for_display_from_hscodes(new_hscodes, release2)
    removed_hscodes_lineitems = get_lineitems_for_display_from_hscodes(removed_hscodes, release1)
    changed1 = get_lineitems_for_display_from_hscodes(hscodes_with_change,release1)
    changed2 = get_lineitems_for_display_from_hscodes(hscodes_with_change,release2)
    changed = [item for pair in zip(changed1, changed2) for item in pair]

    new_lineitems_excel = __convert_lineitems_to_excel(new_hscodes_lineitems, 'NEW')
    removed_lineitems_excel = __convert_lineitems_to_excel(removed_hscodes_lineitems, 'REMOVED')
    changed_excel = __convert_lineitems_to_excel(changed, 'CHANGED')

    return new_lineitems_excel, removed_lineitems_excel, changed_excel
    

def __convert_lineitems_to_excel(lineitems: list, cell_highlight_preset: str) -> BytesIO:
    """Converts a list of line-items to an excel file.
    The cell_highlight_preset specifies how to format the excel file (NEW means everything is highlighted in green)
    (REMOVED means everything is highlighted in red)
    (CHANGED means rows are considered in pairs (old and new), and thick borders separate each pair. Changed cells only are highlighted in blue)

    Args:
        lineitems (list): _description_
        cell_highlight_preset (str): "NEW","REMOVED" or "CHANGED"

    Returns:
        BytesIO: excel file
    """
    if lineitems == None or len(lineitems) == 0:
        excel = BytesIO()
        blank_df = pd.DataFrame()
        blank_df.to_excel(excel, index=False, engine='openpyxl')
        excel.seek(0)
        return excel

    df = pd.DataFrame(lineitems)
    column_order = [
        "Release",
        "Chapter Number",
        "HS Hdg", 
        "HS Code", 
        "SC Code","HS Hdg Name",
        "Prefix",
        "Description", 
        "Unit",
        "ICL/SLSI","Preferential Duty_AP",
        "Preferential Duty_AD",
        "Preferential Duty_BN","Preferential Duty_GT",
        "Preferential Duty_IN",
        "Preferential Duty_PK","Preferential Duty_SA",
        "Preferential Duty_SF",
        "Preferential Duty_SD","Preferential Duty_SG",
        "Gen Duty",
        "VAT",
        "PAL_Gen",
        "PAL_SG","Cess_GEN",
        "Cess_SG",
        "Excise SPD",
        "Surcharge on Customs Duty","SSCL","SCL"]
    df = df[column_order]
    excel = BytesIO()
    df.to_excel(excel, index=False, engine='openpyxl')
    excel.seek(0)

    if cell_highlight_preset == 'NEW':
        wb = oxl.load_workbook(excel)
        ws = wb.active
        highlight_fill = oxlstyles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=ws.min_column, max_col=ws.max_column):
            for cell in row: cell.fill = highlight_fill
        wb.save(excel)
        excel.seek(0)

    if cell_highlight_preset == 'REMOVED':
        wb = oxl.load_workbook(excel)
        ws = wb.active
        highlight_fill = oxlstyles.PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=ws.min_column, max_col=ws.max_column):
            for cell in row: cell.fill = highlight_fill
        wb.save(excel)
        excel.seek(0)

    if cell_highlight_preset == 'CHANGED':
        wb = oxl.load_workbook(excel)
        ws = wb.active
        highlight_fill = oxlstyles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        thick_border = oxlstyles.Border(bottom=oxlstyles.Side(border_style="thick", color="000000"))  # Black thick bottom border

        # Add thick border below the header row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).border = thick_border  # Header row (row 1)

        # Iterate through the rows in pairs
        for row_index in range(2, ws.max_row, 2):  # Start from 2 (skip header row), increment by 2 to handle pairs
            top_row = ws[row_index]  # First row in the pair
            if row_index + 1 <= ws.max_row:  # Ensure there's a second row in the pair
                bottom_row = ws[row_index + 1]

                # Compare cells in the top and bottom row
                for col_index, (top_cell, bottom_cell) in enumerate(zip(top_row, bottom_row), start=1):
                    if top_cell.value != bottom_cell.value:
                        # Highlight the cell in the bottom row
                        ws.cell(row=row_index + 1, column=col_index).fill = highlight_fill

                # Add a thick bottom border below the pair
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_index + 1, column=col).border = thick_border

        wb.save(excel)
        excel.seek(0)

    return excel
